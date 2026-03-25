from django.shortcuts import render, redirect
from django.contrib import messages
import json
import os
from datetime import datetime
import traceback
import tempfile
from django.conf import settings

from django.http import HttpResponse, JsonResponse, FileResponse

from . import  master_json2excel   # the module we created
from .gemini_service import generate_timetable_from_prompt 

def master_generate_page(request):
    return render(request, "master_ai/master_timetable.html")

# ─────────────────────────────────────────────────────────────────────────────
# HELPER – saves timetable dict to generated_timetables/master_timetables.json
# ─────────────────────────────────────────────────────────────────────────────
def _save_master_json(data: dict) -> str:
    """
    Saves `data` as pretty JSON to:
        <BASE_DIR>/generated_timetables/master_timetables.json
    Returns the full file path.
    """
    output_dir  = os.path.join(settings.BASE_DIR, "generated_timetables")
    os.makedirs(output_dir, exist_ok=True)
    file_path   = os.path.join(output_dir, "master_timetables.json")
 
    with open(file_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
 
    print(f"✅ Master timetable JSON saved to: {file_path}")
    return file_path


# ─────────────────────────────────────────────────────────────────────────────
# MASTER TIMETABLE VIEW  –  builds prompt → sends to Gemini → saves JSON
# ─────────────────────────────────────────────────────────────────────────────
def master_timetable_view(request):
    """
    1. Collects form data
    2. Builds the master prompt
    3. Sends prompt to Gemini API
    4. Saves Gemini's JSON response to generated_timetables/master_timetables.json
    5. Redirects to master_result page
    """
    if request.method != "POST":
        return HttpResponse("Only POST method allowed", status=405)
 
    try:
        # ── COMMON ACADEMIC STRUCTURE ────────────────────────────────────
        working_days        = request.POST.get("working_days",       "Mon,Tue,Wed,Thu,Fri")
        start_time          = request.POST.get("start_time",         "09:30")
        end_time            = request.POST.get("end_time",           "17:00")
        lecture_duration    = request.POST.get("lecture_duration",   "60")
        lab_duration        = request.POST.get("lab_duration",       "120")
        break_duration      = request.POST.get("break_duration",     "30")
        break_position      = request.POST.get("break_position",     "fixed")
 
        # ── DYNAMIC CLASSES ──────────────────────────────────────────────
        class_names         = request.POST.getlist("class_name[]")
        subject_json_list   = request.POST.getlist("subject_json[]")
        teacher_json_list   = request.POST.getlist("teacher_json[]")
 
        # ── BATCH & INFRASTRUCTURE ───────────────────────────────────────
        total_students      = request.POST.get("total_students",     "180")
        students_per_batch  = request.POST.get("students_per_batch", "60")
        total_batches       = request.POST.get("total_batches",      "3")
        lab_allocation_mode = request.POST.get("lab_allocation_mode","parallel")
        classrooms = request.POST.getlist("classrooms[]")
        labs = request.POST.getlist("labs[]")
 
        # ── PER-CLASS FIELDS ─────────────────────────────────────────────
        batch_names   = request.POST.getlist("batch_names[]")
        mini_project  = request.POST.getlist("mini_project[]")
 
        total_classes = len(class_names)
 
        # ── Debug ────────────────────────────────────────────────────────
        print("\n" + "="*60)
        print("DEBUG — POST data received")
        print(f"  class_names      : {class_names}")
        print(f"  subject_json_list: {[s[:80]+'…' if len(s)>80 else s for s in subject_json_list]}")
        print(f"  teacher_json_list: {[t[:80]+'…' if len(t)>80 else t for t in teacher_json_list]}")
        print("="*60 + "\n")
 
        if total_classes == 0:
            messages.error(request,
                "No classes submitted. Please generate class sections first.")
            return redirect('master_generate')
 
        # ── BUILD PROMPT ─────────────────────────────────────────────────
        prompt = f"""
══════════════════════════════════════════════════════════════════════════════
MASTER TIMETABLE GENERATION REQUEST
══════════════════════════════════════════════════════════════════════════════
 
College: Your Engineering College
Academic Structure:
- Working Days     : {working_days}
- Daily Start Time : {start_time}
- Daily End Time   : {end_time}
- Lecture Duration : {lecture_duration} Hours
- Lab Duration     : {lab_duration} Hours
- Break Duration   : {break_duration} minutes
- Break Position   : {break_position}
 
Total Classes: {total_classes}
Classes: {', '.join(class_names)}
 
Batch Configuration:
- Total Students     : {total_students}
- Students per Batch : {students_per_batch}
- Total Batches      : {total_batches}
- Lab Mode           : {lab_allocation_mode}
- Classrooms: {classrooms}
- Labs: {labs}
 
══════════════════════════════════════════════════════════════════════════════
CLASS-WISE DATA
══════════════════════════════════════════════════════════════════════════════
"""
 
        for i in range(total_classes):
            class_name         = class_names[i]
            class_batch_names  = batch_names[i]   if i < len(batch_names)  else "N/A"
            class_mini_project = mini_project[i]  if i < len(mini_project) else "N/A"
            subject_json       = subject_json_list[i] if i < len(subject_json_list) else "{}"
            teacher_json       = teacher_json_list[i] if i < len(teacher_json_list) else "{}"
 
            # Validate JSON fields
            try:
                json.loads(subject_json) if subject_json.strip() else None
            except json.JSONDecodeError:
                subject_json = f"[INVALID JSON for class {class_name}]"
 
            try:
                json.loads(teacher_json) if teacher_json.strip() else None
            except json.JSONDecodeError:
                teacher_json = f"[INVALID JSON for class {class_name}]"
 
            prompt += f"""
CLASS {i+1}: {class_name}
──────────────────────────────────────────────
SUBJECT DATA:
{subject_json if subject_json.strip() else "(No subject data uploaded)"}
 
TEACHER DATA:
{teacher_json if teacher_json.strip() else "(No teacher data uploaded)"}
 
Batch Names : {class_batch_names}
 
Special Blocks:
- Mini Project : {class_mini_project} Hours
 
"""
            
        HARD_CONSTRAINTS_PROMPT = """
            ══════════════════════════════════════════════════════════════════════════════
            LAB TIME SLOT ATTACHMENT 
            ══════════════════════════════════════════════════════════════════════════════
            - SE :- allocate all SE lab at 9:30 to 11:30
            - TE :- allocate all TE lab at 11:30 to 1:30
            - BE :- allocate all BE lab at 2:00 to 4:00

            ══════════════════════════════════════════════════════════════════════════════
            HARD CONSTRAINTS  (EVERY RULE MUST BE SATISFIED — ZERO EXCEPTIONS)
            ══════════════════════════════════════════════════════════════════════════════
            
            ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            RULE 1 — NO ROOM DOUBLE-BOOKING  (most critical)
            ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            On any given day, a physical room (e.g. 319, 320, 326) can be occupied by
            AT MOST ONE batch/class during any time window.
            
            For lab slots the room number is embedded in the batch string:
                Format:  SUBJECT-TEACHER-ROOM-BATCH   e.g. "DBMS-KLM-326-S1"
                                                                        ^^^
                                                                    room = 326
            
            Before assigning a room to a lab batch, check: is that room already used
            by another class or batch at an overlapping time? If yes → use a different room.
            
            ❌ WRONG — both SE and TE use room 319 at 9:30-11:30 on WED:
                SE  1_9.30-11.30  row1: "MDML-KLM-319-S1"   ← room 319
                TE  1_9.30-10.30  row1: "IOT-NA-319-T1"      ← room 319 CLASH!
            
            ✅ CORRECT — rooms are distinct:
                SE  1_9.30-11.30  row1: "MDML-KLM-319-S1"   ← room 319
                TE  1_9.30-10.30  row1: "IOT-NA-321-T1"      ← room 321 ✓
            
            ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            RULE 2 — NO TEACHER DOUBLE-BOOKING
            ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            A teacher can teach AT MOST ONE class or batch at any point in time across
            all classes (SE, TE, BE) on a given day.
            
            For theory slots: row2 is the teacher code (e.g. "KLM", "ARS").
            For lab slots:    teacher code is the 2nd segment of the batch string
                            e.g. "DBMS-KLM-326-S1" → teacher is "KLM"
            
            Before assigning a teacher, check all other classes at the same time slot.
            If that teacher is already scheduled elsewhere → use a different teacher.
            
            ❌ WRONG — teacher KLM is teaching both SE and TE at 10:30-11:30 on TUE:
                SE  2_10.30-11.30  row2: "KLM"               ← KLM busy
                TE  2_10.30-11.30  row1: "SPCCL-KLM-319-T1"  ← KLM CLASH!
            
            ✅ CORRECT — different teachers:
                SE  2_10.30-11.30  row2: "KLM"
                TE  2_10.30-11.30  row1: "SPCCL-PSJ-319-T1"  ← PSJ instead ✓
            
            ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            RULE 3 — LAB BATCH ROOM ROTATION (stagger batches across classes)
            ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            When multiple classes have lab sessions at the same time, each batch from each
            class must be placed in a DIFFERENT physical room.
            
            Each class has batches S1/S2/S3 (or T1/T2/T3 or B1/B2/B3).
            If SE has a 3-batch lab at time T using rooms [319, 320, 321],
            then TE's lab at the same time T must use a completely different set of rooms
            (e.g. [322, 323, 325]) — there must be ZERO overlap.
            
            Lab scheduling rule:
            • Divide available lab rooms into groups.
            • Assign Group A to SE labs, Group B to TE labs, Group C to BE labs.
            • Never assign a room from one group to a different class on the same day/slot.
            
            Example room groups (adjust to actual available rooms):
            SE labs  → rooms: 318, 319, 320
            TE labs  → rooms: 321, 322, 323
            BE labs  → rooms: 325, 326, 327
            
            ❌ WRONG — SE and TE labs share room 319 at the same time:
                SE  1_9.30-11.30  row1: "MDML-KLM-319-S1"    ← room 319
                                row2: "DTL-BRP-320-S2"      ← room 320
                                row3: "DBMSL-MD-321-S3"     ← room 321
                TE  1_9.30-10.30  row1: "IOT-NA-319-T1"       ← room 319 CLASH!
            
            ✅ CORRECT — SE uses 318/319/320, TE uses 321/322/323:
                SE  1_9.30-11.30  row1: "MDML-KLM-318-S1"    ← room 318
                                row2: "DTL-BRP-319-S2"      ← room 319
                                row3: "DBMSL-MD-320-S3"     ← room 320
                TE  1_9.30-10.30  row1: "IOT-NA-321-T1"       ← room 321 ✓
            
            ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            RULE 4 — NO CLASS CAN ATTEND TWO SUBJECTS SIMULTANEOUSLY
            ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            A single class (SE, TE, or BE as a whole) cannot have two different subjects
            at the same time. Batch-split labs are the only exception — they are allowed
            because the class is physically split into sub-groups (S1, S2, S3) each in a
            different room.
            
            ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            RULE 5 — MINI PROJECT AND LONG BLOCKS MUST CLEAR ROOM/TEACHER CONFLICTS
            ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            When a class has Mini Project or any multi-slot block spanning 2+ hours:
            • The rooms used by that class during those hours must NOT appear in
                any other class's schedule at the same time.
            • If Mini Project uses a shared lab space, treat it as a reserved room block.
            
            Mini Project slot format — use a SINGLE key spanning the full duration:
                "1_9.30-1.30": {"row1": "MINI PROJECT", "row2": "ALL", "row3": "304"}
            DO NOT repeat the same block under multiple keys for the same class.
            
            ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            RULE 6 — LAB FORMAT AND KEY NAMING
            ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            Lab batch strings MUST follow this exact format:
                SUBJECT_ABBREV-TEACHER_CODE-ROOM_NUMBER-BATCH_ID
                Example: "DBMS-KLM-326-S1"
            
            For a 2-hour lab spanning two time slots, use a SINGLE key with the full range:
                "1_9.30-11.30": {"row1": "DBMS-KLM-318-S1",
                                "row2": "OSL-ARS-319-S2",
                                "row3": "BMDL-BRP-320-S3"}
            DO NOT emit the same lab under two separate slot keys.

            -You have to Give every subject lab to each batch(ex:-S1 , S2 , S3) at once , at different day Different Batch.
            
            ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            RULE 7 — BREAK / RECESS MUST BE IDENTICAL ACROSS ALL CLASSES
            ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            Every class must have:   "RECESS": "RECESS"
            at exactly the same time slot (1.30-2.00). No class may have a lecture or
            lab scheduled during the recess window.
            
            ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            VERIFICATION CHECKLIST — run this mentally before finalising JSON output
            ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            For every day, for every time slot:
            □ List all rooms used across all classes → any duplicates? Fix them.
            □ List all teachers used across all classes → any duplicates? Fix them.
            □ Check lab rooms per class form non-overlapping sets.
            □ Check no class has two entries for the same time window.
            □ Check RECESS is present in every class on every day.
            □ Check multi-hour blocks are expressed as ONE key, not multiple.
══════════════════════════════════════════════════════════════════════════════
HARD CONSTRAINTS (MUST BE STRICTLY FOLLOWED)
══════════════════════════════════════════════════════════════════════════════
1. No teacher can teach two different classes at the same time
2. No room can be used by two classes simultaneously
3. No class can attend two subjects at the same time
4. Lab batches run in parallel but must respect room availability
5. Break must be exactly at the specified time slot
6. Mini Project must be scheduled as per input
 
Return ONLY valid JSON in this exact structure (no extra text, no markdown):
 
{
  "timetable": [
    {
      "day": "MON",
      "classes": [
        {
          "class_name": "SE",
          "schedule": {
            "9.30-10.30": ["Subject", "Teacher(Shortform)", "Room"]
          }
        }
      ]
    }
  ]
}


########### Structue Example #################
{
          "timetable": {
            "slots": [
              "1\n9.30-10.30", 
              "2\n10.30-11.30", 
              "3\n11.30-12.30", 
              "4\n12.30-1.30", 
              "RECESS\n1.30-2.00", 
              "5\n2.00-3.00", 
              "6\n3.00-4.00", 
              
            ],
            "days": [
              {
                "day": "MON",
                "classes": [
                  {
                    "class": "SE",
                    "schedule": {
                      "1_9.30-10.30":  {"row1": "DBMS", "row2": "MD",  "row3": "322"},
                      "2_10.30-11.30": {"row1": "OS",   "row2": "ARS", "row3": "322"},
                      "3_11.30-12.30": {"row1": "DBMS-KLM-326-S1", "row2": "OS-VAB-319-S2", "row3": "BMD-PSJ-320-S3"},
                      "4_12.30-1.30":  {"row1": "DBMS-KLM-326-S1", "row2": "OS-VAB-319-S2", "row3": "BMD-PSJ-320-S3"},
                      "RECESS": "RECESS",
                      "5_2.00-3.00":   {"row1": "MDM",  "row2": "KLM", "row3": "322"}
                    }
                  },

                  {
                    "class": "TE",
                    "schedule": {
                      "1_9.30-10.30":  {"row1": "IoT", "row2": "PSS", "row3": "323"},
                      "2_10.30-11.30": {"row1": "CC",  "row2": "MD",  "row3": "323"},
                      "3_11.30-12.30": {"row1": "AI",  "row2": "MMD", "row3": "323"},
                      "4_12.30-1.30":  {"row1": "SPCC","row2": "PSJ", "row3": "323"},
                      "RECESS": "RECESS",
                      "5_2.00-3.00":  {"row1": "SPCCL-PSJ-319-T1", "row2": "CSSL-SRC-325A-T2", "row3": "MCL-SVT-321-T3"},
                      "6_3.00-4.00":  {"row1": "SPCCL-PSJ-319-T1", "row2": "CSSL-SRC-325A-T2", "row3": "MCL-SVT-321-T3"}
                    }
                  },
                  {
                    "class": "BE",
                    "schedule": {
                      "1_9.30-10.30":  {"row1": "PM",  "row2": "SVT", "row3": "304"},
                      "2_10.30-11.30": {"row1": "SMA", "row2": "SRC", "row3": "304"},
                      "3_11.30-12.30": {"row1": "ADS", "row2": "ARS", "row3": "304"},
                      "4_12.30-1.30":  {"row1": "DC",  "row2": "RCS", "row3": "304"},
                      "RECESS": "RECESS",
                      "5_2.00-3.00":   {"row1": "DCL-RCS-325A-B1", "row2": "ADSL-ARS-326-B2", "row3": "SMAL-BRP-320-B3"},
                      "6_3.00-4.00":   {"row1": "DCL-RCS-325A-B1", "row2": "ADSL-ARS-326-B2", "row3": "-"}
                    }
                  }
                ]
              }
            ]
          }
        }


 
Return Only JSON code.

        """
 
        prompt += HARD_CONSTRAINTS_PROMPT
 
        # ── Print & save prompt ──────────────────────────────────────────
        print("\n" + "="*100)
        print("MASTER TIMETABLE PROMPT")
        print("="*100)
        print(prompt)
        print("="*100 + "\n")
 
        prompt_dir  = os.path.join(settings.BASE_DIR, "generated_timetables", "master_prompts")
        os.makedirs(prompt_dir, exist_ok=True)
        prompt_path = os.path.join(prompt_dir, "master_prompt.txt")
        with open(prompt_path, "w", encoding="utf-8") as f:
            f.write(prompt)
        print(f"✅ Prompt saved to: {prompt_path}")
 
        # ── SEND TO GEMINI ────────────────────────────────────────────────
        print("\n🚀 Sending prompt to Gemini API...")
        messages.info(request, "Sending prompt to Gemini AI — please wait...")
 
        gemini_result = generate_timetable_from_prompt(prompt)
 
        # ── HANDLE GEMINI RESPONSE ────────────────────────────────────────
        if gemini_result["status"] == "error":
            error_msg = gemini_result.get("message", "Unknown Gemini error")
            print(f"❌ Gemini API error: {error_msg}")
            messages.error(request, f"Gemini API error: {error_msg}")
            return redirect('master_generate')
 
        timetable_data = gemini_result["data"]
        print("✅ Gemini response received successfully")
        print(f"   Keys in response: {list(timetable_data.keys()) if isinstance(timetable_data, dict) else type(timetable_data)}")
 
        # ── SAVE JSON TO FILE ─────────────────────────────────────────────
        saved_path = _save_master_json(timetable_data)
 
        messages.success(
            request,
            f"Master timetable generated for {total_classes} class(es) and saved successfully!"
        )
 
        return redirect('master_result')
 
    except Exception as e:
        print("❌ ERROR in master_timetable_view:")
        print(traceback.format_exc())
        messages.error(request, f"Error: {str(e)}")
        return redirect('master_generate')
    


# ─────────────────────────────────────────────────────────────────────────────
# 1.  RESULT PAGE  –  renders master_result.html
#     URL name : 'master_result'
#     Call this view after your prompt is generated and the JSON is ready.
#     Pass the timetable JSON string via session or query param.
# ─────────────────────────────────────────────────────────────────────────────
def master_result_view(request):
    """
    Renders master_result.html.
 
    Reads the timetable JSON from the file:
        <BASE_DIR>/generated_timetables/master_timetables.json
 
    This file is saved by master_timetable_view after the AI generates it.
    """
    json_file_path = os.path.join(
        settings.BASE_DIR,
        "generated_timetables",
        "master_timetables.json"
    )
 
    print(f"[master_result_view] Looking for JSON at: {json_file_path}")
 
    if not os.path.exists(json_file_path):
        return render(request, 'master_ai/master_result.html', {
            'error': (
                f'Timetable file not found at: {json_file_path}. '
                'Please generate the master timetable first.'
            ),
            'timetable_json': '',
        })
 
    try:
        with open(json_file_path, 'r', encoding='utf-8') as f:
            raw = f.read().strip()
 
        # Pretty-print so the textarea is readable
        timetable_json = json.dumps(json.loads(raw), indent=2)
        print(f"[master_result_view] Loaded JSON — {len(timetable_json)} chars")
 
    except json.JSONDecodeError as e:
        return render(request, 'master_ai/master_result.html', {
            'error': f'JSON file is invalid: {e}',
            'timetable_json': '',
        })
    except Exception as e:
        return render(request, 'master_ai/master_result.html', {
            'error': f'Could not read timetable file: {e}',
            'timetable_json': '',
        })
 
    return render(request, 'master_ai/master_result.html', {
        'timetable_json': timetable_json,
    })
 
 
# ─────────────────────────────────────────────────────────────────────────────
# 2.  DOWNLOAD  –  converts JSON → Excel and returns as file download
#     URL name : 'master_json_to_excel'
#     Method   : POST
#     Form field: json_data (the full master timetable JSON string)
# ─────────────────────────────────────────────────────────────────────────────
def master_json_to_excel(request):
    """Convert master timetable JSON → Excel and send as download."""
    if request.method != "POST":
        return HttpResponse("Only POST allowed", status=405)
 
    json_str = request.POST.get('json_data', '').strip()
 
    print("\n" + "=" * 70)
    print("master_json_to_excel — download requested")
    print(f"JSON length : {len(json_str)}")
    print("JSON sample :", json_str[:200] + "…" if len(json_str) > 200 else json_str)
    print("=" * 70 + "\n")
 
    if not json_str:
        messages.error(request, "No JSON data received.")
        return redirect('master_result')
 
    try:
        data = json.loads(json_str)
 
        # Create a temp file for the Excel output
        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        tmp_path = tmp.name
        tmp.close()
 
        # Build the Excel file using our external module
        master_json2excel.build_timetable(data, tmp_path)
 
        print(f"Excel created — size: {os.path.getsize(tmp_path)} bytes")
 
        if not os.path.exists(tmp_path) or os.path.getsize(tmp_path) == 0:
            messages.error(request, "Excel generation failed — file is empty.")
            return redirect('master_result')
 
        response = FileResponse(
            open(tmp_path, 'rb'),
            as_attachment=True,
            filename='master_timetable.xlsx',
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        return response
 
    except json.JSONDecodeError as e:
        print(f"JSON parse error: {e}")
        messages.error(request, f"Invalid JSON: {e}")
        return redirect('master_result')
 
    except Exception as e:
        print("Error in master_json_to_excel:")
        print(traceback.format_exc())
        messages.error(request, f"Error generating Excel: {e}")
        return redirect('master_result')
 
 
# ─────────────────────────────────────────────────────────────────────────────
# 3.  PREVIEW  –  converts JSON → Excel → HTML table (AJAX)
#     URL name : 'master_preview_excel'
#     Method   : POST (AJAX fetch from master_result.html)
#     Returns  : JsonResponse { success, html } or { error }
# ─────────────────────────────────────────────────────────────────────────────
def master_preview_excel(request):
    """Generate an HTML preview of the master timetable Excel (called via AJAX)."""
    if request.method != "POST":
        return JsonResponse({"error": "Only POST allowed"}, status=405)
 
    json_str = request.POST.get('json_data', '').strip()
 
    if not json_str:
        return JsonResponse({"error": "No JSON data received."}, status=400)
 
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
 
        data = json.loads(json_str)
 
        # Build temp Excel
        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        tmp_path = tmp.name
        tmp.close()
 
        master_json2excel.build_timetable(data, tmp_path)
 
        # ── Read Excel and convert to HTML ───────────────────────────────
        wb = load_workbook(tmp_path)
        ws = wb.active
 
        # Map merged ranges: start_coord → {colspan, rowspan}
        merged_map  = {}
        merged_skip = set()
 
        for mr in ws.merged_cells.ranges:
            start = mr.start_cell.coordinate
            merged_map[start] = {
                'colspan': mr.max_col - mr.min_col + 1,
                'rowspan': mr.max_row - mr.min_row + 1,
            }
            for row in range(mr.min_row, mr.max_row + 1):
                for col in range(mr.min_col, mr.max_col + 1):
                    coord = f"{get_column_letter(col)}{row}"
                    if coord != start:
                        merged_skip.add(coord)
 
        html = (
            '<div class="table-responsive mt-3" '
            'style="max-height:600px; overflow:auto;">'
            '<table class="table table-bordered excel-preview-table">'
        )
 
        for row_idx in range(1, ws.max_row + 1):
            html += '<tr>'
            for col_idx in range(1, ws.max_column + 1):
                coord = f"{get_column_letter(col_idx)}{row_idx}"
 
                if coord in merged_skip:
                    continue
 
                cell  = ws.cell(row=row_idx, column=col_idx)
                value = str(cell.value) if cell.value is not None else ""
 
                colspan = merged_map.get(coord, {}).get('colspan', 1)
                rowspan = merged_map.get(coord, {}).get('rowspan', 1)
 
                # ── Cell style ───────────────────────────────────────────
                style = "padding:6px 8px; text-align:center; vertical-align:middle; border:1px solid #999;"
 
                # Header rows (1-2): grey background
                if row_idx <= 2:
                    style += "background-color:#D9D9D9; font-weight:bold; font-size:0.85rem;"
 
                # DAY column (col 1)
                elif col_idx == 1:
                    style += "background-color:#D9D9D9; font-weight:bold;"
 
                # CLASS column (col 2)
                elif col_idx == 2:
                    style += "background-color:#F2F2F2; font-weight:bold;"
 
                # Lab / batch cells (light green)
                elif any(x in value for x in
                         ["-S1","-S2","-S3","-T1","-T2","-T3","-B1","-B2","-B3"]):
                    style += "background-color:#E2EFDA;"
 
                # RECESS cell
                elif value.upper() == "RECESS":
                    style += "font-weight:bold;"
 
                # Normal data cell
                else:
                    style += "background-color:#ffffff;"
 
                # Long values: allow wrap
                if len(value) > 18:
                    style += "white-space:normal; word-break:break-word; max-width:160px;"
 
                html += (
                    f'<td style="{style}" '
                    f'rowspan="{rowspan}" colspan="{colspan}">'
                    f'{value}</td>'
                )
            html += '</tr>'
 
        html += '</table></div>'
 
        os.unlink(tmp_path)
 
        return JsonResponse({"success": True, "html": html})
 
    except json.JSONDecodeError as e:
        return JsonResponse({"error": f"Invalid JSON: {e}"}, status=400)
 
    except Exception as e:
        print("Error in master_preview_excel:")
        print(traceback.format_exc())
        return JsonResponse({"error": str(e)}, status=500)
 