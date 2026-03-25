"""
master_json2excel.py
--------------------
Django-importable module.  Call:

    from master_json2excel import build_timetable
    build_timetable(data_dict, "/tmp/out.xlsx")

Fixes:
  1. Non-standard time-range keys (e.g. "5_2.00-4.00", "1_9.30-11.30")
     are parsed by their START time and END time, then mapped to the correct
     column span automatically — no more silent drops.
  2. Lab/batch blocks where row1≠row2≠row3 are written as 3 stacked merged rows
     (each row spans both columns), matching the reference format.
  3. Multi-slot keys (e.g. "1_9.30-1.30" spanning 4 columns) are fully supported.
  4. MINI PROJECT and FREE blocks render correctly.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Style primitives ─────────────────────────────────────────────────────
THIN   = Side(style="thin",   color="BFBFBF")
MED    = Side(style="medium", color="7F7F7F")
NONE_S = Side(style=None)


def _bdr(left=THIN, right=THIN, top=THIN, bottom=THIN):
    return Border(left=left, right=right, top=top, bottom=bottom)


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _fnt(bold=False, size=9, color="000000"):
    return Font(name="Segoe UI", bold=bold, size=size, color=color)


def _aln(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ── Color palette ────────────────────────────────────────────────────────
HEADER_BG = _fill("2B3A55")   # deep navy
DAY_BG    = _fill("E8EDED")   # soft slate
CLASS_BG  = _fill("F4F6F9")   # very light blue/grey
WHITE_BG  = _fill("FFFFFF")   # theory slots
LAB_BG    = _fill("E2EFDA")   # lab / batch slots
RECESS_BG = _fill("FFF2CC")   # recess
EMPTY_BG  = _fill("FAFAFA")   # empty slots


# ── Column layout ────────────────────────────────────────────────────────
# Each slot header: (slot_number_label, time_label, excel_col_letter)
SLOT_HEADERS = [
    ("1", "9.30-10.30",  "C"),
    ("2", "10.30-11.30", "D"),
    ("3", "11.30-12.30", "E"),
    ("4", "12.30-1.30",  "F"),
    ("",  "1.30-2.00",   "G"),   # RECESS column
    ("5", "2.00-3.00",   "H"),
    ("6", "3.00-4.00",   "I"),
]

# Map: start-time string → Excel column letter
START_TO_COL = {
    "9.30":  "C",
    "10.30": "D",
    "11.30": "E",
    "12.30": "F",
    "2.00":  "H",
    "3.00":  "I",
}

# Map: end-time string → last Excel column letter covered
END_TO_COL = {
    "10.30": "C",   # slot ends at 10.30 → only col C
    "11.30": "D",   # slot ends at 11.30 → up to col D
    "12.30": "E",   # slot ends at 12.30 → up to col E
    "1.30":  "F",   # slot ends at 1.30  → up to col F
    "3.00":  "H",   # slot ends at 3.00  → up to col H
    "4.00":  "I",   # slot ends at 4.00  → up to col I
}

# All data column letters in order (no recess G)
DATA_COLS_ORDER = ["C", "D", "E", "F", "H", "I"]

COL_WIDTHS = {
    "A": 7,  "B": 8,
    "C": 15, "D": 15, "E": 18, "F": 16,
    "G": 8,
    "H": 18, "I": 18,
}

CLASSES    = ["SE", "TE", "BE"]
CLASS_ROWS = 3


# ── Cell writer ──────────────────────────────────────────────────────────
def _ap(ws, cell, value="", bold=False, size=9, font_color="000000",
        fill_=None, h="center", v="center",
        left=THIN, right=THIN, top=THIN, bottom=THIN):
    cell.value     = value
    cell.font      = _fnt(bold=bold, size=size, color=font_color)
    cell.fill      = fill_ if fill_ is not None else WHITE_BG
    cell.alignment = _aln(h=h, v=v)
    cell.border    = _bdr(left=left, right=right, top=top, bottom=bottom)


# ── Slot-key parser ──────────────────────────────────────────────────────
def _parse_key_to_cols(key: str):
    """
    Parse any schedule key and return (start_col, end_col) or None.

    Handles:
      "RECESS"            → None  (handled separately)
      "1_9.30-10.30"      → ("C", "C")
      "5_2.00-4.00"       → ("H", "I")
      "1_9.30-11.30"      → ("C", "D")
      "1_9.30-1.30"       → ("C", "F")
      "2_10.30-11.30"     → ("D", "D")
    """
    if key == "RECESS":
        return None

    # Strip the leading slot-number prefix (e.g. "1_", "5_")
    # The time range is after the first "_"
    if "_" in key:
        # Some keys: "1_9.30-10.30", others: "5_2.00-4.00"
        # Split on first underscore to get the time part
        parts = key.split("_", 1)
        time_part = parts[1]          # e.g. "9.30-10.30" or "9.30-11.30"
    else:
        time_part = key

    # Split time_part on "-" — but times like "9.30-10.30" have multiple dashes
    # Strategy: split on the dash that separates start from end
    # Start time is always before the first dash that precedes a digit group > 9.30
    # Simplest: split by "-" and reconstruct
    dash_parts = time_part.split("-")
    # dash_parts for "9.30-10.30" → ["9.30", "10.30"]
    # dash_parts for "9.30-11.30" → ["9.30", "11.30"]
    # dash_parts for "9.30-1.30"  → ["9.30", "1.30"]
    # dash_parts for "2.00-4.00"  → ["2.00", "4.00"]

    if len(dash_parts) < 2:
        return None

    start_time = dash_parts[0].strip()   # e.g. "9.30"
    end_time   = dash_parts[1].strip()   # e.g. "10.30"

    start_col = START_TO_COL.get(start_time)
    end_col   = END_TO_COL.get(end_time)

    if not start_col or not end_col:
        return None

    return (start_col, end_col)


def _cols_between(start_col: str, end_col: str):
    """Return list of column letters from start_col to end_col (data cols only)."""
    result = []
    in_range = False
    for col in DATA_COLS_ORDER:
        if col == start_col:
            in_range = True
        if in_range:
            result.append(col)
        if col == end_col:
            break
    return result


def _is_lab(r1: str) -> bool:
    if not r1:
        return False
    if "PROJECT" in r1.upper() or "FREE" in r1.upper():
        return True
    return any(x in r1 for x in
               ["-S1", "-S2", "-S3", "-T1", "-T2", "-T3", "-B1", "-B2", "-B3"])


def _slot_vals(slot):
    if not isinstance(slot, dict):
        return "", "", ""
    return slot.get("row1", ""), slot.get("row2", ""), slot.get("row3", "")


# ── Block writer ─────────────────────────────────────────────────────────
def _write_block(ws, start_col, end_col, cls_start, cls_end, bot_cls,
                 r1, r2, r3, cfill):
    """
    Write one logical time block spanning start_col..end_col, rows cls_start..cls_end.

    Cases:
      A. Single column, all rows same  → merge all 3 rows into 1 cell
      B. Single column, rows differ    → 3 separate cells stacked
      C. Multi-column, all rows same   → merge all rows × all cols into 1 cell
      D. Multi-column, rows differ     → merge cols per sub-row (3 wide cells stacked)
    """
    rs        = MED if end_col == "I" else THIN
    all_same  = (r1 == r2 == r3)
    all_empty = (not r1 and not r2 and not r3)
    if all_empty:
        cfill = EMPTY_BG

    if start_col == end_col:
        # ── Single column ────────────────────────────────────────────────
        if all_same:
            ws.merge_cells(f"{start_col}{cls_start}:{start_col}{cls_end}")
            _ap(ws, ws[f"{start_col}{cls_start}"], r1,
                fill_=cfill, left=THIN, right=rs, top=THIN, bottom=bot_cls)
        else:
            for i, (rn, val) in enumerate(
                    zip([cls_start, cls_start+1, cls_start+2], [r1, r2, r3])):
                _ap(ws, ws[f"{start_col}{rn}"], val,
                    bold=(i == 0 and bool(val)),
                    fill_=cfill, left=THIN, right=rs,
                    top=THIN, bottom=bot_cls if i == 2 else NONE_S)
    else:
        # ── Multi-column ─────────────────────────────────────────────────
        if all_same:
            # One big merged cell across all rows and columns
            ws.merge_cells(f"{start_col}{cls_start}:{end_col}{cls_end}")
            _ap(ws, ws[f"{start_col}{cls_start}"], r1,
                fill_=cfill, left=THIN, right=rs, top=THIN, bottom=bot_cls)
        else:
            # One merged cell per sub-row (row spans all columns)
            for i, (rn, val) in enumerate(
                    zip([cls_start, cls_start+1, cls_start+2], [r1, r2, r3])):
                ws.merge_cells(f"{start_col}{rn}:{end_col}{rn}")
                _ap(ws, ws[f"{start_col}{rn}"], val,
                    bold=(i == 0 and bool(val)),
                    fill_=cfill, left=THIN, right=rs,
                    top=THIN, bottom=bot_cls if i == 2 else NONE_S)


# ── Public API ───────────────────────────────────────────────────────────
def build_timetable(data: dict, output_path: str) -> None:
    """
    Build the master timetable Excel and save to output_path.

    Parameters
    ----------
    data        : dict – {"timetable": {"days": [...]}} or {"days": [...]}
    output_path : str  – absolute path for the .xlsx output file
    """
    timetable = data.get("timetable", data)
    days_list = timetable.get("days", [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Master Timetable"
    ws.sheet_view.showGridLines = False

    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[col].width = w

    # ── Header rows 1 & 2 ────────────────────────────────────────────────
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 18

    ws.merge_cells("A1:A2")
    _ap(ws, ws["A1"], "DAY", bold=True, size=10, font_color="FFFFFF",
        fill_=HEADER_BG, left=MED, right=THIN, top=MED, bottom=THIN)

    ws.merge_cells("B1:B2")
    _ap(ws, ws["B1"], "CLASS", bold=True, size=10, font_color="FFFFFF",
        fill_=HEADER_BG, left=THIN, right=MED, top=MED, bottom=THIN)

    for (num, time, col) in SLOT_HEADERS:
        is_rec = (col == "G")
        label  = "RECESS" if is_rec else (f"SLOT {num}" if num else "")
        rs     = MED if col == "I" else THIN
        _ap(ws, ws[f"{col}1"], label, bold=True, size=10, font_color="FFFFFF",
            fill_=HEADER_BG, left=THIN, right=rs, top=MED, bottom=THIN)
        _ap(ws, ws[f"{col}2"], "" if is_rec else time, size=9, font_color="D9D9D9",
            fill_=HEADER_BG, left=THIN, right=rs, top=THIN, bottom=MED)

    # ── Data rows ─────────────────────────────────────────────────────────
    current_row = 3
    days_data   = {
        d["day"]: {c["class"]: c["schedule"] for c in d["classes"]}
        for d in days_list
    }

    for day_info in days_list:
        day       = day_info["day"]
        day_start = current_row
        day_end   = current_row + len(CLASSES) * CLASS_ROWS - 1

        ws.merge_cells(f"A{day_start}:A{day_end}")
        _ap(ws, ws[f"A{day_start}"], day, bold=True, size=11, fill_=DAY_BG,
            left=MED, right=THIN, top=MED, bottom=MED)

        for cls in CLASSES:
            cls_start   = current_row
            cls_end     = current_row + CLASS_ROWS - 1
            is_last_cls = (cls == CLASSES[-1])
            bot_cls     = MED if is_last_cls else THIN

            for r in range(cls_start, cls_end + 1):
                ws.row_dimensions[r].height = 18

            ws.merge_cells(f"B{cls_start}:B{cls_end}")
            _ap(ws, ws[f"B{cls_start}"], cls, bold=True, size=10, fill_=CLASS_BG,
                left=THIN, right=MED, top=THIN, bottom=bot_cls)

            schedule = days_data.get(day, {}).get(cls, {})

            # ── RECESS column — always written first ────────────────────
            ws.merge_cells(f"G{cls_start}:G{cls_end}")
            _ap(ws, ws[f"G{cls_start}"], "RECESS", bold=True,
                size=9, font_color="B45F06", fill_=RECESS_BG,
                left=THIN, right=THIN, top=THIN, bottom=bot_cls)

            # ── Track which data columns are already filled ─────────────
            filled_cols = set()   # column letters already written in this class row

            # ── Process each schedule key ───────────────────────────────
            for key, slot in schedule.items():
                if key == "RECESS":
                    continue   # already written above

                col_range = _parse_key_to_cols(key)
                if col_range is None:
                    print(f"  [WARN] Cannot parse key '{key}' — skipping")
                    continue

                start_col, end_col = col_range

                # Skip if any column in this range is already filled
                span_cols = _cols_between(start_col, end_col)
                if any(c in filled_cols for c in span_cols):
                    print(f"  [WARN] Columns {span_cols} already filled — skipping key '{key}'")
                    continue

                r1, r2, r3 = _slot_vals(slot)
                cfill = LAB_BG if _is_lab(r1) else WHITE_BG

                _write_block(ws, start_col, end_col,
                             cls_start, cls_end, bot_cls,
                             r1, r2, r3, cfill)

                for c in span_cols:
                    filled_cols.add(c)

            # ── Fill any remaining unfilled data columns as empty ───────
            for col in DATA_COLS_ORDER:
                if col not in filled_cols:
                    ws.merge_cells(f"{col}{cls_start}:{col}{cls_end}")
                    _ap(ws, ws[f"{col}{cls_start}"], "",
                        fill_=EMPTY_BG, left=THIN,
                        right=MED if col == "I" else THIN,
                        top=THIN, bottom=bot_cls)

            current_row += CLASS_ROWS

        # ── Enforce thick bottom border across the full day band ─────────
        for (_, _, col) in SLOT_HEADERS:
            c = ws[f"{col}{day_end}"]
            c.border = Border(left=c.border.left, right=c.border.right,
                              top=c.border.top, bottom=MED)
        ws[f"B{day_end}"].border = Border(
            left=ws[f"B{day_end}"].border.left, right=MED,
            top=ws[f"B{day_end}"].border.top,   bottom=MED)

    ws.freeze_panes = "C3"
    wb.save(output_path)
    print(f"[master_json2excel] Saved → {output_path}")