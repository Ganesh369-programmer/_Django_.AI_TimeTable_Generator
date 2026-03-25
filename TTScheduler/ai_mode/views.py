from django.shortcuts import render , redirect 
from django.http import JsonResponse, HttpResponse
from django.template.loader import render_to_string
from django.conf import settings
import pandas as pd
import json
import traceback
import os
from datetime import datetime

from io import BytesIO
from django.http import FileResponse

# Import Gemini service
from .gemini_service import generate_timetable_from_prompt
# from .deepseek_service import generate_timetable_from_prompt
from . import json2excel
from django.contrib import messages
import base64

# =====================================================
# MAIN PAGES
# =====================================================

def generate_page(request):
    return render(request, "ai_mode/generate.html")


def result_page(request):
    """
    Load latest generated timetable JSON file
    and display it on result page
    """
    filename = request.session.get("latest_timetable_file")
    if not filename:
        return render(request, "ai_mode/result.html", {
            "error": "No timetable has been generated yet."
        })

    file_path = os.path.join(settings.BASE_DIR, "generated_timetables", filename)
    
    if not os.path.exists(file_path):
        return render(request, "ai_mode/result.html", {
            "error": f"Timetable file not found: {filename}"
        })

    try:
        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        
        # Prepare list of departments with their JSON for textarea
        departments_list = []
        for key, dept_data in data.items():
            # Make pretty JSON string for display/copy
            pretty_json = json.dumps(dept_data, indent=2, ensure_ascii=False)
            
            # NEW: Pre-format clean timetable JSON as string (double quotes, valid)
            if 'timetable' in dept_data:
                clean_timetable = {"timetable": dept_data['timetable']}
                timetable_json_str = json.dumps(clean_timetable, indent=2)  # Proper JSON string
            else:
                timetable_json_str = '{}'  # Fallback empty

            departments_list.append({
                "name": dept_data.get("name", key.replace("department_", "Department ")),  # nicer title
                "json_pretty": pretty_json,
                "raw_data": dept_data,   # if you still want to render table too
                "timetable_json": timetable_json_str,  # Clean JSON for textarea
            })

        # Optional: Check for generated Excel in session
        has_excel = bool(request.session.get('excel_file') and os.path.exists(request.session.get('excel_file', '')))
        excel_filename = request.session.get('excel_filename', '')

        return render(request, "ai_mode/result.html", {
            "departments": departments_list,
            "filename": filename,          # show which file this is
            "has_excel": has_excel,        # for download button
            "excel_filename": excel_filename,
        })
    
    except Exception as e:
        messages.error(request, f"Error reading timetable: {str(e)}")
        return render(request, "ai_mode/result.html", {
            "error": f"Error reading timetable: {str(e)}"
        })

# =====================================================
# AJAX SUBJECT FORM LOADER
# =====================================================

def get_subject_form(request):
    if request.method == "GET":
        subject = request.GET.get("subject", "")
        if not subject:
            return JsonResponse({"error": "No subject provided"}, status=400)

        html = render_to_string(
            "ai_mode/form_template.html",
            {"subject": subject},
            request=request
        )

        return JsonResponse({"html": html})

    return JsonResponse({"error": "Method not allowed"}, status=405)


# =====================================================
# SAFE HELPER FUNCTION (Prevents IndexError)
# =====================================================

def safe_get(lst, index, default=""):
    try:
        return lst[index]
    except (IndexError, TypeError):
        return default


# =====================================================
# MAIN TIMETABLE GENERATION
# =====================================================

def timetable_form(request):

    if request.method != "POST":
        return HttpResponse("Invalid request method", status=405)

    try:
        # ===============================
        # COLLECT FORM DATA
        # ===============================

        dept_names = request.POST.getlist("dept_name[]")
        semesters = request.POST.getlist("semester[]")
        working_days = request.POST.getlist("working_days[]")
        lecture_duration = request.POST.getlist("lecture_duration[]")
        lab_duration = request.POST.getlist("lab_duration[]")
        break_name = request.POST.getlist("break_name[]")
        classrooms = request.POST.getlist("classrooms[]")
        labs = request.POST.getlist("labs[]")

        start_time = request.POST.getlist("start_time[]")
        end_time = request.POST.getlist("end_time[]")
        break_duration = request.POST.getlist("break_duration[]")
        break_position = request.POST.getlist("break_position[]")

        total_classrooms = request.POST.getlist("total_classrooms[]")
        total_labs = request.POST.getlist("total_labs[]")
        shared_rooms = request.POST.getlist("shared_rooms[]")

        mini_project = request.POST.getlist("mini_project[]")

        total_students = request.POST.getlist("total_students[]")
        students_per_batch = request.POST.getlist("students_per_batch[]")
        total_batches = request.POST.getlist("total_batches[]")
        batch_names = request.POST.getlist("batch_names[]")
        lab_mode = request.POST.getlist("lab_allocation_mode[]")

        subject_files = request.FILES.getlist("subject_excel[]")
        teacher_files = request.FILES.getlist("teacher_excel[]")

        total_departments = len(dept_names)

        if total_departments == 0:
            return render(request, "ai_mode/error.html", {
                "message": "No departments submitted."
            })

        combined_result = {}

        # ===============================
        # LOOP THROUGH EACH DEPARTMENT
        # ===============================

        for i in range(total_departments):

            # -------- Subject Excel → JSON --------
            subject_json = []
            if i < len(subject_files):
                df_subject = pd.read_excel(subject_files[i])
                subject_json = df_subject.to_dict(orient="records")

            # -------- Teacher Excel → JSON --------
            teacher_json = []
            if i < len(teacher_files):
                df_teacher = pd.read_excel(teacher_files[i])
                teacher_json = df_teacher.to_dict(orient="records")

            # -------- Construct AI Prompt SAFELY --------
            example_format = {
                    "department_1": {
                        "name": "Computer 2nd Year",
                        "timetable": [
                        {
                            "day": "MON",
                            "schedule": [
                            { "time": "9.30-10.30", "subject": "AM-3", "class": "SE-COMP", "room": "321" },
                            { "time": "10.30-11.30", "subject": "DSGT", "class": "SE-COMP",  "room": "321" },
                            { "time": "11.30-1.30", "type": "Lab/Practical", "details": ["AI Lab-Deshpande-319-T1", "MC Lab-Tandale-320-T2", "SPCC Lab-Jogdand-318-T3"] },
                            { "time": "1.30-2.00", "type": "Break" }
                            ]
                        },
                        {
                            "day": "TUE",
                            "schedule": [
                            { "time": "9.30-10.30", "subject": "CG", "class": "SE-COMP", "room": "321" },
                            { "time": "10.30-11.30", "subject": "AM-3", "class": "SE-COMP", "room": "321" }
                            ]
                        }
                        ]
                    },

                    "department_2": {
                        "name": "Computer 3rd Year",
                        "timetable": [
                        {
                            "day": "MON",
                            "schedule": [
                            { "time": "9.30-11.30", "type": "Lab/Practical", "details": ["Mini Project-Staff-322-All"] },
                            { "time": "11.30-12.30", "subject": "AI", "class": "TE-COMP", "room": "322" }
                            ]
                        }
                        ]
                    }
                    }

            prompt = f"""
==============================
DEPARTMENT {i+1}
==============================
Department: {safe_get(dept_names,i)}
Semester: {safe_get(semesters,i)}

Working Days: {safe_get(working_days,i)}
Lecture Duration: {safe_get(lecture_duration,i)} Hour
Lab Duration: {safe_get(lab_duration,i)} Hour
Break: {safe_get(break_name,i)} Minutes

Classrooms: {safe_get(classrooms,i)}
Labs: {safe_get(labs,i)}

SUBJECT DATA:
{json.dumps(subject_json, indent=2)}

TEACHER DATA:
{json.dumps(teacher_json, indent=2)}

 Academic Structure:
Start Time: {safe_get(start_time,i)}
End Time: {safe_get(end_time,i)}
Break Duration: {safe_get(break_duration,i)}
Break Position: {safe_get(break_position,i)}

Infrastructure:
Total Classrooms: {safe_get(total_classrooms,i)}
Total Labs: {safe_get(total_labs,i)}
Shared Rooms: {safe_get(shared_rooms,i)}

Special Blocks:
Mini Project: {safe_get(mini_project,i)}

Batch System:
Total Students: {safe_get(total_students,i)}
Students Per Batch: {safe_get(students_per_batch,i)}
Total Batches: {safe_get(total_batches,i)}
Batch Names: {safe_get(batch_names,i)}
Lab Mode: {safe_get(lab_mode,i)}


**HARD CONSTRAINTS (MUST NEVER BE VIOLATED):**

1. **NO TEACHER CONFLICTS**: A teacher can teach only ONE class at a time
   - Check: Teacher "Deshpande" cannot be in Room 319 AND Room 322 at the same time
   - Example CONFLICT: Monday 9:30-10:30 - "Deshpande teaching AI in 322" AND "Deshpande teaching SPCC in 318" ❌

2. **NO ROOM CONFLICTS**: A room can host only ONE class/lab at a time
   - Check: Room 322 cannot have "TE-COMP Theory" AND "SE-COMP Lab" simultaneously in Between multiple Department
   - Example CONFLICT: Monday 10:00-11:00 - "Room 322: AI Theory" AND "Room 322: MC Lab in between Department" ❌

3. **NO CLASS CONFLICTS**: A class/batch cannot be in TWO places at the same time
   - Check: "TE-COMP" students cannot attend AI lecture AND MC lab simultaneously
   - For lab batches: T1, T2, T3 are DIFFERENT batches, so they CAN be scheduled together
   - Example CONFLICT: "TE-COMP attending AI in Room 322" AND "TE-COMP Lab in Room 318" ❌
   - Give different Time slotes for labs for differenet Department .

4. **TIME SLOT INTEGRITY**: 
   - Theory classes: exactly 1 hour (9:30-10:30, 10:30-11:30, etc.)
   - Labs: exactly 2 hours (9:30-11:30, 11:30-1:30, 2:00-4:00)
   - Break: 30 minutes (1:30-2:00)
   - NO overlapping time slots

### STRICT JSON OUTPUT FORMAT REQUIRED:
                Follow this structure EXACTLY. 
                - For Theory: {{"time": "9.30-10.30", "subject": "Name", "class": "ShortCode", "room": "No"}}
                - For Miniproject: {{ "time": "9.30-11.30", "type": "Lab/Practical", "details": ["Mini Project-Staff-322-All"] }}
                - For Labs: {{"time": "11.30-1.30", "type": "Lab/Practical", "details": ["SUB-TEACHER-ROOM-BATCH"]}}
                - For Breaks: {{"time": "1.30-2.00", "type": "Break"}}

### This is Example
{json.dumps(example_format, indent=2)}

Return STRICT JSON only in this format:
"""
            print(prompt)
            # === SAVE PROMPT TO TEXT FILE INSTEAD OF PRINTING ===
            # === REPLACE MODE - Only latest prompt stays in the file ===
            # === NEW CODE: Create fresh file with time-date + ALL prompts ===
            output_folder = "generated_timetables/prompt"
            os.makedirs(output_folder, exist_ok=True)
            # First collect ALL old department prompts (so "all" are added)
            combined_prompts = ""

            # Find every previous department prompt file
            prompt_files = sorted([f for f in os.listdir(output_folder) 
                                if f.startswith("department_") and f.endswith("_prompt.txt")])

            for file_name in prompt_files:
                with open(os.path.join(output_folder, file_name), "r", encoding="utf-8") as f:
                    combined_prompts += f.read() + "\n\n" + "=" * 60 + "\n\n"

            # Now add the current new prompt
            combined_prompts += "=" * 60 + "\n"
            combined_prompts += prompt
            combined_prompts += "\n" + "=" * 60 + "\n"

            # Create NEW file name with current time & date
            from datetime import datetime                     # ← add this line at top of file if not present
            timestamp = datetime.now().strftime("%d-%m-%Y_%H-%M-%S")
            text_file_name = f"timetable_prompt_{timestamp}.txt"
            file_path = os.path.join(output_folder, text_file_name)

            # Write everything to the NEW file
            with open(file_path, "w", encoding="utf-8") as file:
                file.write(combined_prompts)

            print(f"✅ New combined file created: {file_path}")



            # -------- Call Gemini --------
            result = generate_timetable_from_prompt(prompt)

            if result["status"] == "success":
                raw = result["data"].strip() if isinstance(result["data"], str) else str(result["data"])

                print(f"\n{'─'*80}")
                print("RAW AI RESPONSE:")
                print(f"{'─'*80}")
                print(raw[:500] + "..." if len(raw) > 500 else raw)
                print(f"{'─'*80}\n")

                # ========================================
                # ENHANCED CLEANING & PARSING
                # ========================================
                
                # Remove markdown code blocks
                if raw.startswith("```json"):
                    raw = raw.split("```json", 1)[-1]
                if raw.startswith("```"):
                    raw = raw.split("```", 1)[-1]
                if raw.endswith("```"):
                    raw = raw.rsplit("```", 1)[0]
                raw = raw.strip()

                # Try parsing as JSON first
                parsed_json = None
                try:
                    parsed_json = json.loads(raw)
                    print("✅ Successfully parsed as valid JSON")
                
                except json.JSONDecodeError as json_err:
                    # print(f"⚠️ JSON parse failed: {json_err}")
                    # print("🔄 Attempting to fix Python dict format (single quotes → double quotes)...")
                    
                    # Convert Python dict string to JSON
                    try:
                        import ast
                        import re
                        
                        # Method 1: Use ast.literal_eval (safe evaluation)
                        try:
                            parsed_json = ast.literal_eval(raw)
                            print("✅ Parsed using ast.literal_eval")
                        except:
                            # Method 2: Manual regex replacement
                            # Replace single quotes with double quotes (carefully)
                            fixed = raw.replace("'", '"')
                            # Fix common issues
                            fixed = re.sub(r'True', 'true', fixed)
                            fixed = re.sub(r'False', 'false', fixed)
                            fixed = re.sub(r'None', 'null', fixed)
                            
                            parsed_json = json.loads(fixed)
                            print("✅ Parsed after quote conversion")
                    
                    except Exception as fix_err:
                        print(f"❌ All parsing attempts failed: {fix_err}")
                        parsed_json = None

                # ========================================
                # STORE RESULT
                # ========================================
                
                if parsed_json:
                    # Extract the correct structure
                    if "name" in parsed_json and "timetable" in parsed_json:
                        # Direct format: {"name": "...", "timetable": [...]}
                        combined_result[f"department_{i+1}"] = parsed_json
                        print(f"✅ Stored department_{i+1} with {len(parsed_json.get('timetable', []))} days")
                    
                    elif f"department_{i+1}" in parsed_json:
                        # Nested format: {"department_1": {"name": "...", "timetable": [...]}}
                        combined_result[f"department_{i+1}"] = parsed_json[f"department_{i+1}"]
                        print(f"✅ Extracted nested department_{i+1}")
                    
                    else:
                        # Fallback: assume it's timetable array
                        combined_result[f"department_{i+1}"] = {
                            "name": safe_get(dept_names, i, f"Department {i+1}"),
                            "timetable": parsed_json if isinstance(parsed_json, list) else []
                        }
                        print(f"⚠️ Used fallback structure for department_{i+1}")
                
                else:
                    # Parsing failed completely
                    combined_result[f"department_{i+1}"] = {
                        "error": "Failed to parse AI response",
                        "raw_output": raw[:1000]  # Store first 1000 chars for debugging
                    }
                    print(f"❌ Stored error for department_{i+1}")

            else:
                # AI service error
                combined_result[f"department_{i+1}"] = {
                    "error": result["message"]
                }
                print(f"❌ AI service error: {result['message']}")

        # ===============================
        # SAVE JSON TO FILE
        # ===============================
        filename = "latest_timetables.json"           # ← fixed name

        folder_path = os.path.join(settings.BASE_DIR, "generated_timetables")
        os.makedirs(folder_path, exist_ok=True)
        file_path = os.path.join(folder_path, filename)

        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(combined_result, f, indent=4)

        # Store in session (still useful, but now it's always the same name)
        request.session["latest_timetable_file"] = filename
        # ✅ REDIRECT instead of JSON
        return redirect("/ai_mode/result")

    # At the end of except block in timetable_form()
    except Exception as e:
        print("❌ ERROR OCCURRED:")
        print(traceback.format_exc())
        # Option A: render error page (current)
        # return render(request, "ai_mode/error.html", {"message": str(e)})

        # Option B: better UX – redirect back with message
        from django.contrib import messages
        messages.error(request, f"Timetable generation failed: {str(e)}")
        return redirect("generate") 
    

import time
import tempfile
def json_2_excel_converter(request):
    """Convert JSON to Excel and provide download"""
    if request.method != "POST":
        return HttpResponse("Only POST allowed", status=405)

    json_str = request.POST.get('json_data', '').strip()
    dept_name = request.POST.get('dept_name', 'unknown')

    print("\n" + "="*80)
    print(f"Converting for: {dept_name}")
    print(f"JSON length: {len(json_str)}")
    print("JSON sample:", json_str[:200] + "..." if len(json_str) > 200 else json_str)
    print("="*80 + "\n")

    if not json_str:
        print("ERROR: No JSON received")
        messages.error(request, "No JSON data received.")
        return redirect('result')

    try:
        full_data = json.loads(json_str)
        print("Parsed data keys:", list(full_data.keys()))

        # Extract timetable
        if "timetable" in full_data:
            data = {"timetable": full_data["timetable"]}
            print(f"Timetable extracted: {len(data['timetable'])} days")
        else:
            data = full_data
            print("No 'timetable' key — using full data")

        # Create temp file
        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        temp_path = temp_file.name
        temp_file.close()

        print(f"Calling build_timetable with path: {temp_path}")

        # Generate Excel
        json2excel.build_timetable(data, temp_path)

        print(f"After build_timetable — file exists? {os.path.exists(temp_path)}")
        print(f"File size: {os.path.getsize(temp_path) if os.path.exists(temp_path) else 0} bytes")

        if os.path.exists(temp_path) and os.path.getsize(temp_path) > 0:
            # Store in session for later download/preview
            request.session[f'excel_file_{dept_name}'] = temp_path
            request.session[f'excel_filename_{dept_name}'] = f"{dept_name.replace(' ', '_')}_timetable.xlsx"
            
            # Direct download
            response = FileResponse(
                open(temp_path, 'rb'),
                as_attachment=True,
                filename=f"{dept_name.replace(' ', '_')}_timetable.xlsx",
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            return response
        else:
            print("ERROR: File not created or empty")
            if os.path.exists(temp_path):
                os.unlink(temp_path)
            messages.error(request, "Excel generation failed — check terminal for details.")
            return redirect('result')

    except json.JSONDecodeError as e:
        print(f"JSON error: {e}")
        messages.error(request, f"Invalid JSON: {str(e)}")
        return redirect('result')
    except Exception as e:
        print(f"General error: {e}")
        print("Traceback:", traceback.format_exc())
        messages.error(request, f"Error: {str(e)}")
        return redirect('result')


def preview_excel(request):
    """Generate Excel preview as HTML table with proper structure"""
    if request.method != "POST":
        return JsonResponse({"error": "Only POST allowed"}, status=405)

    json_str = request.POST.get('json_data', '').strip()
    dept_name = request.POST.get('dept_name', 'unknown')

    if not json_str:
        return JsonResponse({"error": "No JSON data received"}, status=400)

    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        
        full_data = json.loads(json_str)
        
        # Extract timetable
        if "timetable" in full_data:
            data = {"timetable": full_data["timetable"]}
        else:
            data = full_data

        # Create temp Excel file
        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        temp_path = temp_file.name
        temp_file.close()

        # Generate Excel
        json2excel.build_timetable(data, temp_path)

        # Read Excel and convert to HTML
        wb = load_workbook(temp_path)
        ws = wb.active

        # Track which cells are part of merged ranges (to skip them)
        merged_cells_coords = set()
        merged_ranges_map = {}  # Maps start cell to its range info
        
        for merged_range in ws.merged_cells.ranges:
            start_cell = merged_range.start_cell.coordinate
            merged_ranges_map[start_cell] = {
                'colspan': merged_range.max_col - merged_range.min_col + 1,
                'rowspan': merged_range.max_row - merged_range.min_row + 1
            }
            # Add all cells in this range to skip list
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    coord = f"{get_column_letter(col)}{row}"
                    if coord != start_cell:
                        merged_cells_coords.add(coord)

        # Build HTML table
        html = '''
        <div class="table-responsive mt-3" style="max-height: 600px; overflow-y: auto;">
            <table class="table table-bordered excel-preview-table" style="font-size: 0.85rem; border-collapse: collapse;">
        '''
        
        for row_idx in range(1, ws.max_row + 1):
            html += '<tr>'
            
            for col_idx in range(1, ws.max_column + 1):
                coord = f"{get_column_letter(col_idx)}{row_idx}"
                
                # Skip if this cell is part of a merged range (but not the start)
                if coord in merged_cells_coords:
                    continue
                
                cell = ws.cell(row=row_idx, column=col_idx)
                value = str(cell.value) if cell.value is not None else ""
                
                # Get merge info if this is a merged cell start
                colspan = 1
                rowspan = 1
                if coord in merged_ranges_map:
                    colspan = merged_ranges_map[coord]['colspan']
                    rowspan = merged_ranges_map[coord]['rowspan']
                
                # Determine cell styling
                style = "padding: 8px; text-align: center; vertical-align: middle; border: 1px solid #dee2e6;"
                
                # Row 1: Period numbers (header row)
                if row_idx == 1:
                    style += "background-color: #2c3e50; color: white; font-weight: bold; font-size: 0.9rem;"
                
                # Row 2: Time slots (header row)
                elif row_idx == 2:
                    style += "background-color: #34495e; color: white; font-weight: bold; font-size: 0.85rem;"
                
                # Column 1: DAY labels
                elif col_idx == 1:
                    style += "background-color: #6c757d; color: white; font-weight: bold; min-width: 60px;"
                
                # Column 2: CLASS/SUBJECT/ROOM labels
                elif col_idx == 2:
                    style += "background-color: #495057; color: #f8f9fa; font-weight: bold; min-width: 80px;"
                
                # Check for BREAK cells
                if "BREAK" in value.upper():
                    style += "background-color: #ffc107; color: #000; font-weight: bold;"
                
                # Check for empty cells in data area
                elif row_idx > 2 and col_idx > 2 and not value.strip():
                    style += "background-color: #f8f9fa;"
                
                # Normal data cells
                elif row_idx > 2 and col_idx > 2:
                    style += "background-color: white; color: #212529;"
                
                # Add word wrapping for lab details
                if "-" in value and len(value) > 20:
                    style += "white-space: normal; word-wrap: break-word; max-width: 200px;"
                
                html += f'<td style="{style}" rowspan="{rowspan}" colspan="{colspan}">{value}</td>'
            
            html += '</tr>'
        
        html += '</table></div>'

        # Clean up temp file
        os.unlink(temp_path)

        return JsonResponse({
            "success": True,
            "html": html,
            "dept_name": dept_name
        })

    except Exception as e:
        print(f"Preview error: {e}")
        print("Traceback:", traceback.format_exc())
        return JsonResponse({
            "error": f"Preview generation failed: {str(e)}"
        }, status=500)


def download_excel(request):
    """Download the generated Excel file"""
    filepath = request.session.get('excel_file')
    filename = request.session.get('excel_filename', 'timetable.xlsx')
    
    if filepath and os.path.exists(filepath):
        response = FileResponse(
            open(filepath, 'rb'), 
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    else:
        messages.error(request, "Excel file not found. Please generate it first.")
        return redirect('result')