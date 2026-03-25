"""
Dynamic Timetable Generator
============================
Rules for the data dictionary:
  • "type": "Break"          → included as a column (no period number, empty header)
  • "type": "Lab/Practical"  → 2-hr merged rows; 'details' list supplies 3 row values
  • No "type" key            → standard theory slot  (subject / class / room rows)
  • Multi-hour, no type      → full 3-row block merge (e.g. MINI PROJECT)

Time format: college shorthand — hours 1-8 mean PM (13:00-20:00).
  "9.30"=09:30  "12.30"=12:30  "1.30"=13:30  "2.00"=14:00  "4.00"=16:00
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter
from openpyxl.cell import MergedCell


# ─────────────────────────────────────────────────────────────────────────────
#  TIME UTILITIES
# ─────────────────────────────────────────────────────────────────────────────

def to_min(t: str) -> int:
    """
    College shorthand → minutes since midnight.
    Hours 1-8 → PM (add 12).  Hours 9-12 → AM (keep as-is).
    """
    t = t.strip()
    if "." in t:
        h, m = int(t.split(".")[0]), int(t.split(".")[1])
    else:
        h, m = int(t), 0
    if h <= 8:          # 1.30 = 13:30, 4.00 = 16:00 etc.
        h += 12
    return h * 60 + m


def split_range(time_str: str):
    """'9.30-10.30' → ('9.30','10.30'),  '11.30-1.30' → ('11.30','1.30')"""
    parts = time_str.strip().split("-")
    return parts[0].strip(), parts[-1].strip()


def min_to_college(minutes: int) -> str:
    """
    Convert minutes-since-midnight back to college shorthand string.
    840 → '2.00',  810 → '1.30',  570 → '9.30'
    """
    h, m = divmod(minutes, 60)
    # Reverse PM conversion
    if h > 12:
        h -= 12
    if m == 0:
        return f"{h}.00"
    return f"{h}.{m:02d}"


def slot_key(s: str, e: str) -> str:
    return f"{s}-{e}"


# ─────────────────────────────────────────────────────────────────────────────
#  INDEX BUILDER
#  Returns:
#    all_slots    – ordered (start, end) for EVERY column (incl. break cols)
#    time_to_col  – slot_key → Excel column number
#    break_set    – set of (start, end) that are break slots
# ─────────────────────────────────────────────────────────────────────────────

def build_index(data: dict):
    all_slots_set = set()   # every slot that needs a column
    break_set     = set()

    for day in data["timetable"]:
        for item in day["schedule"]:
            s, e = split_range(item["time"])
            dur  = to_min(e) - to_min(s)

            if item.get("type") == "Break":
                # Breaks get a column too (shown empty, no period number)
                all_slots_set.add((s, e))
                break_set.add((s, e))

            elif dur <= 60:
                # Single theory hour
                all_slots_set.add((s, e))

            else:
                # Multi-hour span → decompose into 1-hr atomic columns
                cur = to_min(s)
                end = to_min(e)
                while cur < end:
                    nxt  = cur + 60
                    s_hr = min_to_college(cur)
                    e_hr = min_to_college(nxt)
                    all_slots_set.add((s_hr, e_hr))
                    cur  = nxt

    # Sort chronologically by start time
    sorted_slots = sorted(all_slots_set, key=lambda x: to_min(x[0]))

    # Assign Excel columns: col1=DAY, col2=CLASS, slots from col3
    time_to_col = {slot_key(s, e): 3 + i for i, (s, e) in enumerate(sorted_slots)}

    return sorted_slots, time_to_col, break_set


def get_span_cols(s: str, e: str, sorted_slots, time_to_col, break_set):
    """
    Return (first_col, last_col) that a multi-hour span covers.
    Excludes break slots from the span.
    """
    s_min, e_min = to_min(s), to_min(e)
    cols = []
    for a, b in sorted_slots:
        if (a, b) in break_set:
            continue   # don't include break columns in span merges
        if to_min(a) >= s_min and to_min(b) <= e_min and slot_key(a, b) in time_to_col:
            cols.append(time_to_col[slot_key(a, b)])
    return (min(cols), max(cols)) if cols else (None, None)


# ─────────────────────────────────────────────────────────────────────────────
#  SAFE OPENPYXL HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def write_cell(ws, row, col, value=None, font=None, alignment=None, border=None):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return
    if value     is not None: cell.value     = value
    if font      is not None: cell.font      = font
    if alignment is not None: cell.alignment = alignment
    if border    is not None: cell.border    = border


def do_merge(ws, r1, c1, r2, c2):
    try:
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    except Exception:
        pass


def set_border(ws, row, col, border):
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        cell.border = border


# ─────────────────────────────────────────────────────────────────────────────
#  MAIN BUILDER - This is the main function to call from Django
# ─────────────────────────────────────────────────────────────────────────────

def build_timetable(data: dict, output_path: str):
    """
    Main function to generate Excel timetable from JSON data.
    
    Args:
        data (dict): JSON data with 'timetable' key containing schedule
        output_path (str): Full path where Excel file should be saved
    
    Returns:
        None (saves file to output_path)
    """
    
    sorted_slots, time_to_col, break_set = build_index(data)
    total_cols = 2 + len(sorted_slots)

    wb = Workbook()
    ws = wb.active
    ws.title = "Timetable"

    # Styles  (no background fills — exactly like the reference)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin   = Side(style="thin")
    bdr    = Border(left=thin, right=thin, top=thin, bottom=thin)
    f_bold = Font(bold=True)
    f_norm = Font(bold=False)

    def is_break(s, e): return (s, e) in break_set

    # ── ROW 1: Period numbers ─────────────────────────────────────────────────
    write_cell(ws, 1, 1, border=bdr)
    write_cell(ws, 1, 2, border=bdr)
    period = 1
    for s, e in sorted_slots:
        col = time_to_col[slot_key(s, e)]
        if is_break(s, e):
            write_cell(ws, 1, col, border=bdr)              # empty, just border
        else:
            write_cell(ws, 1, col, str(period),
                       font=f_bold, alignment=center, border=bdr)
            period += 1

    # ── ROW 2: Time-slot header labels ────────────────────────────────────────
    write_cell(ws, 2, 1, "DAY",   font=f_bold, alignment=center, border=bdr)
    write_cell(ws, 2, 2, "CLASS", font=f_bold, alignment=center, border=bdr)
    for s, e in sorted_slots:
        col = time_to_col[slot_key(s, e)]
        write_cell(ws, 2, col, slot_key(s, e),
                   font=f_bold, alignment=center, border=bdr)

    # ── COLUMN WIDTHS ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 12
    for s, e in sorted_slots:
        col    = time_to_col[slot_key(s, e)]
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = 13 if is_break(s, e) else 18

    # ── DATA ROWS ─────────────────────────────────────────────────────────────
    cur_row = 3

    for day_data in data["timetable"]:
        day = day_data["day"]

        # DAY label: merged across 3 sub-rows in column A
        do_merge(ws, cur_row, 1, cur_row + 2, 1)
        write_cell(ws, cur_row, 1, day, font=f_bold, alignment=center, border=bdr)

        # Sub-row labels in column B
        for i, lbl in enumerate(["SUBJECT", "CLASS", "ROOM"]):
            write_cell(ws, cur_row + i, 2, lbl,
                       font=f_norm, alignment=center, border=bdr)

        # Pre-fill all columns as empty with border
        for s, e in sorted_slots:
            col = time_to_col[slot_key(s, e)]
            for i in range(3):
                write_cell(ws, cur_row + i, col,
                           font=f_norm, alignment=center, border=bdr)

        # Fill schedule items
        for item in day_data["schedule"]:
            s, e = split_range(item["time"])
            dur  = to_min(e) - to_min(s)

            if item.get("type") == "Break":
                continue    # already rendered as empty bordered cell

            if dur <= 60:
                # ── Theory slot ───────────────────────────────────────────────
                key = slot_key(s, e)
                if key not in time_to_col:
                    print(f"  WARNING: slot '{key}' not in index — skipped")
                    continue
                col = time_to_col[key]
                for i, v in enumerate([item.get("subject", ""),
                                        item.get("class",   ""),
                                        item.get("room",    "")]):
                    write_cell(ws, cur_row + i, col, v,
                               font=f_norm, alignment=center, border=bdr)

            else:
                # ── Multi-hour span ───────────────────────────────────────────
                c1, c2 = get_span_cols(s, e, sorted_slots, time_to_col, break_set)
                if c1 is None:
                    print(f"  WARNING: span '{slot_key(s,e)}' could not be resolved — skipped")
                    continue

                if item.get("type") == "Lab/Practical":
                    # Each sub-row merged across the span; one detail per row
                    details = item.get("details", [])
                    for i in range(3):
                        do_merge(ws, cur_row + i, c1, cur_row + i, c2)
                        v = details[i] if i < len(details) else ""
                        write_cell(ws, cur_row + i, c1, v,
                                   font=f_norm, alignment=center, border=bdr)
                else:
                    # Full 3-row × N-col block (MINI PROJECT etc.)
                    do_merge(ws, cur_row, c1, cur_row + 2, c2)
                    write_cell(ws, cur_row, c1, item.get("subject", ""),
                               font=f_norm, alignment=center, border=bdr)

        # Borders across entire day block
        for r in range(cur_row, cur_row + 3):
            for col in range(1, total_cols + 1):
                set_border(ws, r, col, bdr)

        cur_row += 3

    wb.save(output_path)

    print(f"\n✓  Saved → {output_path}")
    print(f"   Days         : {len(data['timetable'])}")
    print(f"   Total columns: {total_cols}  (2 fixed + {len(sorted_slots)} time slots)")
    print(f"   Break columns: {[slot_key(s,e) for s,e in sorted_slots if is_break(s,e)]}")
    print(f"   Slot order   : {[slot_key(s,e) for s,e in sorted_slots]}")


# ─────────────────────────────────────────────────────────────────────────────
# Optional: Test function if you want to run this file standalone
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    # Test data
    test_data = {
        "timetable": [
            {
                'day': 'MON', 
                'schedule': [
                    {'time': '9.30-11.30', 'type': 'Lab/Practical', 'details': ['Mini Project-Staff-322-All']},
                    {'time': '11.30-12.30', 'subject': 'AI', 'class': 'TE-COMP', 'room': '322'},
                    {'time': '12.30-1.30', 'subject': 'MC', 'class': 'TE-COMP', 'room': '322'},
                    {'time': '1.30-2.00', 'type': 'Break'},
                    {'time': '2.00-4.00', 'type': 'Lab/Practical', 'details': ['AI Lab-Deshpande-319-T1', 'MC Lab-Tandale-320-T2', 'SPCC Lab-Jogdand-318-T3']}
                ]
            }
        ]
    }
    
    build_timetable(test_data, "test_timetable.xlsx")